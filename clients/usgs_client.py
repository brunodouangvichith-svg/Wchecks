"""
Parsing des rapports USGS "Minerals Yearbook, Volume III : Area Reports"
(gratuit, PAS d'API REST — fichiers Excel régionaux téléchargés manuellement depuis
https://www.usgs.gov/centers/national-minerals-information-center et déposés dans
data/usgs/).

LIMITE / TOLÉRANCE DEMANDÉE (voir plan de développement) : le format de ces rapports
varie d'une édition à l'autre et d'une région à l'autre — nom de feuille ("Table 3",
"T3", "T3-Africa", "T3-EUR"...), nombre et ordre des matières premières en colonnes,
présence ou non de sous-catégories (primaire/secondaire). Le parsing est donc
générique plutôt que basé sur des positions de colonnes fixes :
1. la feuille de production est repérée par son nom (T3*/Table 3) ou, à défaut, par
   le texte "PRODUCTION OF SELECTED MINERAL" dans ses premières lignes ;
2. l'en-tête de chaque colonne est reconstitué en concaténant le texte des quelques
   lignes fusionnées au-dessus de la ligne "Country or locality" (le nom de la
   matière première y est réparti sur 2 à 4 lignes) ;
3. chaque colonne est rattachée à une matière première de config.STRATEGIC_MINERALS
   par mots-clés, pas par position — une colonne qui ne correspond à aucun mot-clé
   connu est simplement ignorée.
Les lignes non reconnues (pays introuvable dans country_mapping, valeur non
numérique) sont logguées et ignorées plutôt que de faire planter le parsing.

LIMITE SUPPLÉMENTAIRE CONSTATÉE À L'USAGE : chaque feuille "Table 3" empile en
réalité PLUSIEURS sections (une table "—Continued" plus loin dans la même feuille
réutilise les mêmes colonnes pour d'AUTRES matières premières, avec son propre
en-tête). Le parsing s'arrête volontairement à la fin de la PREMIÈRE section
(voir _is_section_end_marker) pour éviter d'attribuer des valeurs à la mauvaise
matière première. Conséquence concrète observée : le lithium et l'uranium
(mentionnés dans les notes de bas de page de plusieurs fichiers) vivent dans une
section ultérieure non couverte ici — seules les matières présentes dans la
première section de chaque fichier (copper/gold/cobalt selon les régions testées)
sont donc effectivement remontées. Documenté aussi dans le README.
"""

import logging
import re
from pathlib import Path

import openpyxl

from mapping.country_mapping import COUNTRY_NAME_TO_ISO3

logger = logging.getLogger(__name__)

# mot-clé (en minuscule) -> clé matière première (doit correspondre à config.STRATEGIC_MINERALS)
MINERAL_KEYWORDS = {
    "copper": "copper",
    "lithium": "lithium",
    "uranium": "uranium",
    "rare earth": "rare_earths",
    "rare-earth": "rare_earths",
    "gold": "gold",
    "cobalt": "cobalt",
}

NON_VALUES = {"--", "na", "n/a", "w", "nan", ""}

HEADER_LOOKBACK_ROWS = 6  # nb de lignes fusionnées au-dessus de "Country or locality" à concaténer


def _find_production_sheet(wb):
    for name in wb.sheetnames:
        normalized = name.upper().replace(" ", "").replace("-", "")
        if normalized.startswith("T3") or normalized == "TABLE3":
            return wb[name]
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            if any(cell and "PRODUCTION OF SELECTED MINERAL" in str(cell).upper() for cell in row):
                return ws
    raise ValueError("Feuille de production introuvable (aucune feuille 'T3*'/'Table 3' ni titre correspondant).")


def _extract_year(ws) -> int | None:
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        for cell in row:
            if not cell:
                continue
            match = re.search(r"\bIN (\d{4})", str(cell).upper())
            if match:
                return int(match.group(1))
    return None


def _find_header_end_row(ws) -> int:
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        cell = row[0]
        if cell and str(cell).strip().lower().startswith("country"):
            return row_idx
    raise ValueError("Ligne d'en-tête 'Country or locality' introuvable.")


def _build_column_labels(ws, header_end_row: int) -> dict[int, str]:
    start = max(1, header_end_row - HEADER_LOOKBACK_ROWS)
    labels: dict[int, str] = {}
    for row in ws.iter_rows(min_row=start, max_row=header_end_row, values_only=True):
        for col_idx in range(2, len(row), 2):  # colonnes paires = valeurs ; impaires = drapeaux (e/r/...)
            cell = row[col_idx]
            if cell:
                labels[col_idx] = (labels.get(col_idx, "") + " " + str(cell).strip()).strip()
    return labels


def _match_mineral(label: str) -> str | None:
    label_lower = label.lower()
    for keyword, mineral in MINERAL_KEYWORDS.items():
        if keyword in label_lower:
            return mineral
    return None


SECTION_END_MARKERS = (
    "total",  # "Regional total", "World total", "Total" seul
    "share of world",
    "see footnotes",
    "table 3",  # "TABLE 3—Continued"
    "production of selected mineral",  # titre répété = nouvelle section
)


def _is_section_end_marker(cell_text: str) -> bool:
    lowered = cell_text.strip().lower()
    return any(marker in lowered for marker in SECTION_END_MARKERS)


def parse_yearbook_file(path: Path) -> list[dict]:
    """
    Parse un fichier Excel "Minerals Yearbook Volume III" et retourne les lignes
    correspondant aux matières premières de config.STRATEGIC_MINERALS.

    Retourne une liste de dicts {pays_code, annee, matiere_premiere, volume_tonnes}.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = _find_production_sheet(wb)

    year = _extract_year(ws)
    if year is None:
        logger.warning("usgs_client: année introuvable dans '%s', fichier ignoré", path.name)
        return []

    header_end_row = _find_header_end_row(ws)
    column_labels = _build_column_labels(ws, header_end_row)
    column_minerals = {
        col_idx: mineral
        for col_idx, label in column_labels.items()
        if (mineral := _match_mineral(label)) is not None
    }

    if not column_minerals:
        logger.info("usgs_client: aucune matière premère suivie trouvée dans '%s'", path.name)
        return []

    results: list[dict] = []
    unmatched_countries: set[str] = set()

    for row in ws.iter_rows(min_row=header_end_row + 1, values_only=True):
        country_name = row[0]
        if not country_name or not str(country_name).strip():
            continue
        country_name = str(country_name).strip()

        # Ces feuilles empilent plusieurs sections (une table "—Continued" plus loin
        # réutilise les mêmes colonnes pour d'autres matières, avec un nouvel en-tête).
        # On s'arrête à la fin de la PREMIÈRE section pour ne pas attribuer des valeurs
        # au mauvais en-tête — voir limite documentée en tête de module.
        if _is_section_end_marker(country_name):
            break

        pays_code = COUNTRY_NAME_TO_ISO3.get(country_name)
        if pays_code is None:
            unmatched_countries.add(country_name)
            continue

        for col_idx, mineral in column_minerals.items():
            if col_idx >= len(row):
                continue
            raw_value = row[col_idx]
            if raw_value is None or str(raw_value).strip().lower() in NON_VALUES:
                continue
            try:
                value = float(raw_value)
            except (TypeError, ValueError):
                logger.warning(
                    "usgs_client: valeur non numérique ignorée ('%s', %s, %d) : '%s'",
                    country_name, mineral, year, raw_value,
                )
                continue
            results.append(
                {"pays_code": pays_code, "annee": year, "matiere_premiere": mineral, "volume_tonnes": value}
            )

    if unmatched_countries:
        logger.warning(
            "usgs_client: %d pays/localité(s) non reconnu(s) dans '%s', ignoré(s) : %s",
            len(unmatched_countries), path.name, sorted(unmatched_countries),
        )

    logger.info(
        "parse_yearbook_file('%s') : %d ligne(s) extraite(s) (année %d, matières : %s)",
        path.name, len(results), year, sorted(set(column_minerals.values())),
    )
    return results
